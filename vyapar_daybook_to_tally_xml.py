"""
Vyapar Day Book Excel -> Tally Import XML

Purpose:
- Read Vyapar Day Book Excel export
- Generate Tally XML for:
  - New customer ledgers under Sundry Debtors
  - New supplier/expense ledgers under Sundry Creditors / Indirect Expenses
  - New stock items and units if Item Details sheet is present
  - Sales vouchers
  - Receipt vouchers for Money In
  - Payment vouchers for Money Out
  - Purchase vouchers if purchase rows exist
  - Credit Note / Debit Note if rows exist

Usage:
    python vyapar_daybook_to_tally_xml.py --input "Daybook.xlsx" --output "tally_import.xml"

Optional:
    python vyapar_daybook_to_tally_xml.py --input "Daybook.xlsx" --output "tally_import.xml" --company "கொங்கு மாடர்ன் ரைஸ் மில்"

Important:
- Keep Cash, Bank, Sales, Purchase, GST ledgers already created in Tally.
- Configure ledger names below before running.
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

# -----------------------------
# CONFIG - EDIT THESE VALUES
# -----------------------------
COMPANY_NAME = "கொங்கு மாடர்ன் ரைஸ் மில்"

LEDGERS = {
    "sales": "Sales",
    "purchase": "Purchase",
    "cash": "Cash",
    "bank": "TMB - CC 50005",
    "gpay": "TMB - CC 50005",
    "upi": "TMB - CC 50005",
    "phonepe": "TMB - CC 50005",
    "card": "TMB - CC 50005",
    "round_off": "Round Off",
}

PARENTS = {
    "customer": "Sundry Debtors",
    "supplier": "Sundry Creditors",
    "expense": "Indirect Expenses",
    "stock_item": "Primary",
}

# These ledgers will never be auto-created by this script.
BLOCKED_AUTO_CREATE_LEDGERS = {
    "Cash",
    "GPAY",
    "UPI",
    "PhonePe",
    "Google Pay",
    "TMB - CC 50005",
    "Bank Account",
    "Sales",
    "Purchase",
    "CGST",
    "SGST",
    "IGST",
    "Round Off",
    "Mixed Payment Clearing",
}

TYPE_KEYWORDS = {
    "sales": ["pos sale", "sale", "sales"],
    "purchase": ["purchase", "purchase bill"],
    "receipt": ["payment in", "payment-in", "receipt", "money in"],
    "payment": ["payment out", "payment-out", "expense", "money out", "payment"],
    "credit_note": ["credit note", "sale return", "sales return"],
    "debit_note": ["debit note", "purchase return"],
}

# -----------------------------
# LOGGING
# -----------------------------
def setup_logger(log_dir: Path) -> logging.Logger:
    log_dir.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("vyapar_daybook_to_tally")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fmt = logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s")

    file_handler = logging.FileHandler(log_dir / "daybook_to_tally.log", encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(fmt)

    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(fmt)

    logger.addHandler(file_handler)
    logger.addHandler(console)
    return logger

LOGGER = setup_logger(Path("logs"))

ROUND_OFF_WARNING_THRESHOLD = 1.0

# -----------------------------
# DATA MODELS
# -----------------------------
@dataclass
class DaybookRow:
    date: str
    party: str
    ref_no: str
    row_type: str
    payment_type: str
    total: float
    money_in: float
    money_out: float
    description: str
    kind: str

@dataclass
class ItemRow:
    date: str
    ref_no: str
    party: str
    item_name: str
    item_code: str
    hsn: str
    category: str
    quantity: float
    unit: str
    unit_price: float
    discount: float
    tax_percent: float
    tax: float
    transaction_type: str
    amount: float

# -----------------------------
# HELPERS
# -----------------------------
def clean_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def clean_text(value) -> str:
    return str(value or "").strip()


def clean_amount(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    text = text.replace(",", "").replace("₹", "").replace("Rs.", "").replace("Rs", "")
    text = text.replace("CR", "").replace("Dr", "").replace("DR", "").strip()
    if text == "" or text == "-":
        return 0.0
    try:
        return float(text)
    except ValueError:
        LOGGER.warning("Could not parse amount: %r", value)
        return 0.0


def tally_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if value is None:
        return datetime.today().strftime("%Y%m%d")
    text = str(value).strip()
    if not text:
        return datetime.today().strftime("%Y%m%d")
    for fmt in ("%Y%m%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    LOGGER.warning("Could not parse date %r; using today", value)
    return datetime.today().strftime("%Y%m%d")


def classify_type(row_type: str) -> str:
    text = clean_header(row_type)
    for kind, keywords in TYPE_KEYWORDS.items():
        if any(k in text for k in keywords):
            # Avoid classifying payment-in as payment because 'payment' matches first
            if kind == "payment" and ("payment in" in text or "payment-in" in text):
                continue
            return kind
    return "unknown"


def payment_ledger(payment_type: str) -> str:
    text = clean_header(payment_type)
    if "cash" in text and any(x in text for x in ["gpay", "upi", "phonepe", "google pay"]):
        # Vyapar daybook usually does not give split amount. Route to bank ledger or change if needed.
        return LEDGERS["bank"]
    if "cash" in text:
        return LEDGERS["cash"]
    if "gpay" in text or "google pay" in text:
        return LEDGERS["gpay"]
    if "upi" in text:
        return LEDGERS["upi"]
    if "phonepe" in text:
        return LEDGERS["phonepe"]
    if "card" in text:
        return LEDGERS["card"]
    if "bank" in text or "neft" in text or "rtgs" in text or "imps" in text:
        return LEDGERS["bank"]
    return LEDGERS["cash"]


def add_text(parent: ET.Element, tag: str, text) -> ET.Element:
    node = ET.SubElement(parent, tag)
    node.text = "" if text is None else str(text)
    return node


def make_envelope(report_name: str = "Vouchers") -> Tuple[ET.Element, ET.Element]:
    envelope = ET.Element("ENVELOPE")
    header = ET.SubElement(envelope, "HEADER")
    add_text(header, "TALLYREQUEST", "Import Data")
    body = ET.SubElement(envelope, "BODY")
    import_data = ET.SubElement(body, "IMPORTDATA")
    request_desc = ET.SubElement(import_data, "REQUESTDESC")
    add_text(request_desc, "REPORTNAME", report_name)
    static_vars = ET.SubElement(request_desc, "STATICVARIABLES")
    add_text(static_vars, "SVCURRENTCOMPANY", COMPANY_NAME)
    request_data = ET.SubElement(import_data, "REQUESTDATA")
    return envelope, request_data


def write_xml(envelope: ET.Element, path: Path) -> None:
    tree = ET.ElementTree(envelope)
    ET.indent(tree, space="  ")
    path.parent.mkdir(parents=True, exist_ok=True)
    tree.write(path, encoding="utf-8", xml_declaration=True)

# -----------------------------
# EXCEL PARSING
# -----------------------------
def find_sheet_and_header(wb, required_headers: Iterable[str]) -> Tuple[object, int, Dict[str, int]]:
    required = {clean_header(x) for x in required_headers}
    for ws in wb.worksheets:
        for row_no in range(1, min(ws.max_row, 20) + 1):
            headers = {}
            values = []
            for col in range(1, ws.max_column + 1):
                h = clean_header(ws.cell(row_no, col).value)
                if h:
                    headers[h] = col
                    values.append(h)
            if required.issubset(set(values)):
                return ws, row_no, headers
    raise ValueError(f"Could not find sheet/header with required columns: {required_headers}")

def build_date_lookup(item_rows):
    lookup = {}

    for item in item_rows:
        if item.ref_no and item.date:
            lookup[item.ref_no] = item.date

    return lookup


def default_daybook_date(date_by_ref_no: Dict[str, str]) -> str:
    dates = sorted({date for date in date_by_ref_no.values() if date})
    if dates:
        return dates[0]
    fallback = tally_date(None)
    LOGGER.warning("No item dates found; using %s for daybook rows without dates", fallback)
    return fallback
    

def read_daybook(path: Path , date_by_ref_no=None) -> List[DaybookRow]:
    LOGGER.info("Reading Daybook: %s", path)
    date_by_ref_no = date_by_ref_no or {}
    fallback_date = default_daybook_date(date_by_ref_no)
    wb = load_workbook(path, data_only=True)
    ws, header_row, headers = find_sheet_and_header(wb, ["name", "ref no", "type", "payment type", "total", "money in", "money out"])
    LOGGER.info("Daybook sheet: %s | header row: %s", ws.title, header_row)
    LOGGER.debug("Daybook headers: %s", headers)

    rows: List[DaybookRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_type = clean_text(ws.cell(r, headers["type"]).value)
        party = clean_text(ws.cell(r, headers["name"]).value)
        ref_no = clean_text(ws.cell(r, headers["ref no"]).value)
        date_value = date_by_ref_no.get(ref_no) or fallback_date
        ##date=tally_date(ws.cell(r, headers.get("date", 1)).value)
        if not row_type and not party and not ref_no:
            continue
        kind = classify_type(row_type)
        row = DaybookRow(date_value,
            party=party,
            ref_no=ref_no,
            row_type=row_type,
            payment_type=clean_text(ws.cell(r, headers["payment type"]).value),
            total=abs(clean_amount(ws.cell(r, headers["total"]).value)),
            money_in=abs(clean_amount(ws.cell(r, headers["money in"]).value)),
            money_out=abs(clean_amount(ws.cell(r, headers["money out"]).value)),
            description=clean_text(ws.cell(r, headers.get("description", 0)).value) if "description" in headers else "",
            kind=kind,
        )
        rows.append(row)
        LOGGER.debug("Parsed daybook row %s: %s", r, row)
    LOGGER.info("Parsed daybook rows: %s", len(rows))
    return rows


def read_items(path: Path) -> List[ItemRow]:
    LOGGER.info("Reading Item Details, if present: %s", path)
    wb = load_workbook(path, data_only=True)
    try:
        ws, header_row, headers = find_sheet_and_header(wb, ["invoice no./txn no.", "party name", "item name", "quantity", "unit", "amount"])
    except ValueError:
        LOGGER.warning("No Item Details sheet found. Sales/purchase vouchers will be accounting-only, not inventory vouchers.")
        return []

    LOGGER.info("Item Details sheet: %s | header row: %s", ws.title, header_row)
    LOGGER.debug("Item Details headers: %s", headers)
    rows: List[ItemRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        ref_no = clean_text(ws.cell(r, headers["invoice no./txn no."]).value)
        item_name = clean_text(ws.cell(r, headers["item name"]).value)
        if not ref_no or not item_name:
            continue
        row = ItemRow(
            date=tally_date(ws.cell(r, headers.get("date", 1)).value),
            ref_no=ref_no,
            party=clean_text(ws.cell(r, headers["party name"]).value),
            item_name=item_name,
            item_code=clean_text(ws.cell(r, headers.get("item code", 0)).value) if "item code" in headers else "",
            hsn=clean_text(ws.cell(r, headers.get("hsn/sac", 0)).value) if "hsn/sac" in headers else "",
            category=clean_text(ws.cell(r, headers.get("category", 0)).value) if "category" in headers else "",
            quantity=abs(clean_amount(ws.cell(r, headers["quantity"]).value)),
            unit=clean_text(ws.cell(r, headers["unit"]).value) or "UNT",
            unit_price=abs(clean_amount(ws.cell(r, headers.get("unitprice", 0)).value)) if "unitprice" in headers else 0.0,
            discount=abs(clean_amount(ws.cell(r, headers.get("discount", 0)).value)) if "discount" in headers else 0.0,
            tax_percent=abs(clean_amount(ws.cell(r, headers.get("tax percent", 0)).value)) if "tax percent" in headers else 0.0,
            tax=abs(clean_amount(ws.cell(r, headers.get("tax", 0)).value)) if "tax" in headers else 0.0,
            transaction_type=clean_text(ws.cell(r, headers.get("transaction type", 0)).value) if "transaction type" in headers else "",
            amount=abs(clean_amount(ws.cell(r, headers["amount"]).value)),
        )
        rows.append(row)
        LOGGER.debug("Parsed item row %s: %s", r, row)
    LOGGER.info("Parsed item rows: %s", len(rows))
    return rows

# -----------------------------
# MASTER GENERATION
# -----------------------------
def create_ledger_message(request_data: ET.Element, name: str, parent: str, billwise: str = "Yes") -> None:
    if not name or name in BLOCKED_AUTO_CREATE_LEDGERS:
        LOGGER.info("Skipping blocked/system ledger auto-create: %s", name)
        return
    msg = ET.SubElement(request_data, "TALLYMESSAGE")
    ledger = ET.SubElement(msg, "LEDGER", {"NAME": name, "ACTION": "Create"})
    add_text(ledger, "NAME", name)
    add_text(ledger, "PARENT", parent)
    add_text(ledger, "ISBILLWISEON", billwise)
    add_text(ledger, "ISCOSTCENTRESON", "No")
    add_text(ledger, "AFFECTSSTOCK", "No")


def create_unit_message(request_data: ET.Element, unit: str) -> None:
    if not unit:
        return
    msg = ET.SubElement(request_data, "TALLYMESSAGE")
    unit_el = ET.SubElement(msg, "UNIT", {"NAME": unit, "ACTION": "Create"})
    add_text(unit_el, "NAME", unit)
    add_text(unit_el, "ISSIMPLEUNIT", "Yes")
    add_text(unit_el, "DECIMALPLACES", " 2")


def create_stock_item_message(request_data: ET.Element, item: ItemRow) -> None:
    msg = ET.SubElement(request_data, "TALLYMESSAGE")
    stock = ET.SubElement(msg, "STOCKITEM", {"NAME": item.item_name, "ACTION": "Create"})
    add_text(stock, "NAME", item.item_name)
    add_text(stock, "PARENT", PARENTS["stock_item"])
    add_text(stock, "BASEUNITS", item.unit or "UNT")
    add_text(stock, "GSTAPPLICABLE", "Applicable")
    add_text(stock, "GSTTYPEOFSUPPLY", "Goods")
    if item.hsn:
        gst = ET.SubElement(stock, "GSTDETAILS.LIST")
        add_text(gst, "APPLICABLEFROM", datetime.today().strftime("%Y%m%d"))
        add_text(gst, "HSNCODE", item.hsn)
        add_text(gst, "TAXABILITY", "Exempt" if item.tax_percent == 0 else "Taxable")


def build_masters(rows: List[DaybookRow], items: List[ItemRow], output: Path) -> Tuple[ET.Element, Dict[str, int]]:
    envelope, request_data = make_envelope("All Masters")

    sales_parties = {r.party for r in rows if r.kind in {"sales", "receipt", "credit_note"} and r.party}
    purchase_parties = {r.party for r in rows if r.kind in {"purchase", "debit_note"} and r.party}
    expense_parties = {r.party for r in rows if r.kind == "payment" and r.party}
    units = {i.unit for i in items if i.unit}
    stock_by_name: Dict[str, ItemRow] = {}
    for i in items:
        stock_by_name.setdefault(i.item_name, i)

    for party in sorted(sales_parties):
        create_ledger_message(request_data, party, PARENTS["customer"], "Yes")
    for party in sorted(purchase_parties):
        create_ledger_message(request_data, party, PARENTS["supplier"], "Yes")
    # Expenses are optional; create as Indirect Expenses only when row is an expense/payment.
    for party in sorted(expense_parties):
        if party not in sales_parties and party not in purchase_parties:
            create_ledger_message(request_data, party, PARENTS["expense"], "No")
    for unit in sorted(units):
        create_unit_message(request_data, unit)
    for item in sorted(stock_by_name.values(), key=lambda x: x.item_name):
        create_stock_item_message(request_data, item)

    stats = {
        "customer_ledgers": len(sales_parties),
        "supplier_ledgers": len(purchase_parties),
        "expense_ledgers": len(expense_parties),
        "units": len(units),
        "stock_items": len(stock_by_name),
    }
    write_xml(envelope, output)
    return envelope, stats

# -----------------------------
# VOUCHER GENERATION
# -----------------------------
def add_bill_allocation(parent: ET.Element, ref_no: str, amount: float, bill_type: str) -> None:
    bill = ET.SubElement(parent, "BILLALLOCATIONS.LIST")
    add_text(bill, "NAME", ref_no)
    add_text(bill, "BILLTYPE", bill_type)
    add_text(bill, "AMOUNT", f"{amount:.2f}")


def ledger_entry(parent: ET.Element, ledger_name: str, amount: float, is_deemed_positive: str, is_party: str = "No", bill_ref: Optional[str] = None, bill_type: Optional[str] = None,tag_name: str = "ALLLEDGERENTRIES.LIST") -> None:
    entry = ET.SubElement(parent, tag_name)
    add_text(entry, "LEDGERNAME", ledger_name)
    add_text(entry, "ISDEEMEDPOSITIVE", is_deemed_positive)
    add_text(entry, "ISPARTYLEDGER", is_party)
    add_text(entry, "AMOUNT", f"{amount:.2f}")
    if bill_ref and bill_type:
        add_bill_allocation(entry, bill_ref, amount, bill_type)


def add_round_off_entry(voucher: ET.Element, amount: float, voucher_no: str) -> None:
    amount = round(amount, 2)
    if abs(amount) < 0.01:
        return
    if abs(amount) >= ROUND_OFF_WARNING_THRESHOLD:
        LOGGER.warning("Large sales round-off adjustment | voucher=%s | amount=%.2f", voucher_no, amount)
    ledger_entry(
        voucher,
        LEDGERS["round_off"],
        amount,
        "Yes" if amount < 0 else "No",
        tag_name="LEDGERENTRIES.LIST",
    )


def inventory_entry(parent: ET.Element, item: ItemRow, sales_ledger: str, is_sales: bool = True) -> None:
    inv = ET.SubElement(parent, "ALLINVENTORYENTRIES.LIST")
    amount = item.amount if is_sales else -item.amount
    qty_text = f" {item.quantity:g} {item.unit}"
    add_text(inv, "STOCKITEMNAME", item.item_name)
    add_text(inv, "ISDEEMEDPOSITIVE", "No" if is_sales else "Yes")
    add_text(inv, "RATE", f"{item.unit_price:g}/{item.unit}")
    add_text(inv, "DISCOUNT", f"{item.discount:g}")
    add_text(inv, "AMOUNT", f"{amount:.2f}")
    add_text(inv, "ACTUALQTY", qty_text)
    add_text(inv, "BILLEDQTY", qty_text)

    batch = ET.SubElement(inv, "BATCHALLOCATIONS.LIST")
    add_text(batch, "GODOWNNAME", "Main Location")
    add_text(batch, "BATCHNAME", "Primary Batch")
    add_text(batch, "DESTINATIONGODOWNNAME", "Main Location")
    add_text(batch, "AMOUNT", f"{amount:.2f}")
    add_text(batch, "ACTUALQTY", qty_text)
    add_text(batch, "BILLEDQTY", qty_text)

    alloc = ET.SubElement(inv, "ACCOUNTINGALLOCATIONS.LIST")
    add_text(alloc, "LEDGERNAME", sales_ledger)
    add_text(alloc, "ISDEEMEDPOSITIVE", "No" if is_sales else "Yes")
    add_text(alloc, "AMOUNT", f"{amount:.2f}")


def create_voucher(request_data: ET.Element, vchtype: str, date: str, number: str, party: str, reference: str = "") -> ET.Element:
    msg = ET.SubElement(request_data, "TALLYMESSAGE")
    if vchtype =="Sales":
        voucher = ET.SubElement(msg, "VOUCHER", {"VCHTYPE": vchtype, "ACTION": "Create","OBJVIEW": "Invoice Voucher View"})
    else:
        voucher = ET.SubElement(msg, "VOUCHER", {"VCHTYPE": vchtype, "ACTION": "Create"})
    add_text(voucher, "DATE", date)
    add_text(voucher, "VOUCHERTYPENAME", vchtype)
    add_text(voucher, "VOUCHERNUMBER", number)
    add_text(voucher, "REFERENCE", reference or number)
    add_text(voucher, "PARTYNAME", party)
    add_text(voucher, "PARTYLEDGERNAME", party)
    return voucher


def create_sales(request_data: ET.Element, row: DaybookRow, items_by_ref: Dict[str, List[ItemRow]]) -> None:
    voucher = create_voucher(request_data, "Sales", row.date, row.ref_no, row.party, row.ref_no)
    add_text(voucher, "PERSISTEDVIEW", "Invoice Voucher View")
    add_text(voucher, "ISINVOICE", "Yes")
    amount = row.total
    ledger_entry(voucher, row.party, -amount, "Yes", is_party="Yes", bill_ref=row.ref_no, bill_type="New Ref",tag_name="LEDGERENTRIES.LIST")
    items = items_by_ref.get(row.ref_no, [])
    if items:
        item_total = 0.0
        for item in items:
            item_total += item.amount
            inventory_entry(voucher, item, LEDGERS["sales"], is_sales=True)
        add_round_off_entry(voucher, amount - item_total, row.ref_no)
    else:
        ledger_entry(voucher, LEDGERS["sales"], amount, "No")


def create_receipt(request_data: ET.Element, row: DaybookRow, amount: Optional[float] = None, prefix: str = "RCPT") -> None:
    amt = row.money_in if amount is None else amount
    if amt <= 0:
        return
    no = f"{prefix}-{row.ref_no or row.party[:12]}".replace(" ", "-")
    voucher = create_voucher(request_data, "Receipt", row.date, no, row.party, row.ref_no or no)
    bank_or_cash = payment_ledger(row.payment_type)
    # Debit cash/bank and credit the party for receipts.
    ledger_entry(voucher, bank_or_cash, -amt, "Yes")
    ledger_entry(voucher, row.party, amt, "No", "Yes", row.ref_no or no, "Agst Ref" if row.ref_no else "New Ref")


def create_purchase(request_data: ET.Element, row: DaybookRow, items_by_ref: Dict[str, List[ItemRow]]) -> None:
    number = row.ref_no or f"PUR-{row.party[:12]}-{row.date}"
    voucher = create_voucher(request_data, "Purchase", row.date, number, row.party, number)
    add_text(voucher, "PERSISTEDVIEW", "Invoice Voucher View")
    add_text(voucher, "ISINVOICE", "Yes")
    amount = row.total
    ledger_entry(voucher, row.party, amount, "No", "Yes", number, "New Ref")
    items = items_by_ref.get(row.ref_no, [])
    if items:
        for item in items:
            inventory_entry(voucher, item, LEDGERS["purchase"], is_sales=False)
    else:
        ledger_entry(voucher, LEDGERS["purchase"], -amount, "Yes")


def create_payment(request_data: ET.Element, row: DaybookRow, amount: Optional[float] = None, prefix: str = "PYMT") -> None:
    amt = row.money_out if amount is None else amount
    if amt <= 0:
        return
    no = f"{prefix}-{row.ref_no or row.party[:12]}-{row.date}".replace(" ", "-")
    voucher = create_voucher(request_data, "Payment", row.date, no, row.party, row.ref_no or no)
    bank_or_cash = payment_ledger(row.payment_type)
    # Debit expense/supplier, credit cash/bank
    ledger_entry(voucher, row.party, amt, "No", "Yes" if row.ref_no else "No", row.ref_no or no, "Agst Ref" if row.ref_no else "New Ref")
    ledger_entry(voucher, bank_or_cash, -amt, "Yes")


def create_credit_note(request_data: ET.Element, row: DaybookRow) -> None:
    no = row.ref_no or f"CN-{row.party[:12]}-{row.date}"
    voucher = create_voucher(request_data, "Credit Note", row.date, no, row.party, no)
    ledger_entry(voucher, LEDGERS["sales"], -row.total, "Yes")
    ledger_entry(voucher, row.party, row.total, "No", "Yes", no, "New Ref")


def create_debit_note(request_data: ET.Element, row: DaybookRow) -> None:
    no = row.ref_no or f"DN-{row.party[:12]}-{row.date}"
    voucher = create_voucher(request_data, "Debit Note", row.date, no, row.party, no)
    ledger_entry(voucher, row.party, -row.total, "Yes", "Yes", no, "New Ref")
    ledger_entry(voucher, LEDGERS["purchase"], row.total, "No")


def build_vouchers(rows: List[DaybookRow], items: List[ItemRow], output: Path) -> Tuple[ET.Element, Dict[str, int]]:
    envelope, request_data = make_envelope("Vouchers")
    items_by_ref: Dict[str, List[ItemRow]] = defaultdict(list)
    for i in items:
        items_by_ref[i.ref_no].append(i)

    stats = defaultdict(int)
    skipped = []
    for row in rows:
        if row.kind == "sales":
            if not row.ref_no:
                skipped.append((row, "Sales row has no ref no"))
                continue
            create_sales(request_data, row, items_by_ref)
            stats["sales"] += 1
            if row.money_in > 0:
                create_receipt(request_data, row)
                stats["receipts_from_sales"] += 1
        elif row.kind == "purchase":
            create_purchase(request_data, row, items_by_ref)
            stats["purchase"] += 1
            if row.money_out > 0:
                create_payment(request_data, row)
                stats["payments_from_purchase"] += 1
        elif row.kind == "receipt":
            create_receipt(request_data, row, prefix="RCPT-DIRECT")
            stats["direct_receipts"] += 1
        elif row.kind == "payment":
            create_payment(request_data, row, prefix="PYMT-DIRECT")
            stats["direct_payments"] += 1
        elif row.kind == "credit_note":
            create_credit_note(request_data, row)
            stats["credit_notes"] += 1
        elif row.kind == "debit_note":
            create_debit_note(request_data, row)
            stats["debit_notes"] += 1
        else:
            skipped.append((row, "Unknown row type"))

    for row, reason in skipped:
        LOGGER.warning("Skipped row | reason=%s | row=%s", reason, row)
    stats["skipped"] = len(skipped)
    write_xml(envelope, output)
    return envelope, dict(stats)


def combine_xml(master_env: ET.Element, voucher_env: ET.Element, output: Path) -> None:
    # Safer to keep separate, but combined file can be useful for testing.
    envelope, request_data = make_envelope("Vouchers")
    for env in (master_env, voucher_env):
        for msg in env.findall(".//TALLYMESSAGE"):
            request_data.append(msg)
    write_xml(envelope, output)

# -----------------------------
# MAIN
# -----------------------------
def main() -> None:
    global COMPANY_NAME
        
    parser = argparse.ArgumentParser(description="Generate Tally import XML from Vyapar Day Book Excel")
    parser.add_argument("--input", required=True, help="Vyapar Daybook Excel file path")
    parser.add_argument("--output", default="output/final_combined.xml", help="Combined Tally XML output path")
    parser.add_argument("--company", default=COMPANY_NAME, help="Tally company name")
    parser.add_argument("--output-dir", default="output", help="Output folder")
    args = parser.parse_args()
    COMPANY_NAME = args.company
   
    input_path = Path(args.input)
    output_dir = Path(args.output_dir)
    masters_path = output_dir / "01_masters.xml"
    vouchers_path = output_dir / "02_vouchers.xml"
    combined_path = Path(args.output)

    LOGGER.info("Starting Vyapar Daybook to Tally XML")
    LOGGER.info("Input: %s", input_path)
    LOGGER.info("Company: %s", COMPANY_NAME)

    
    
    items = read_items(input_path)
    date_by_ref_no = build_date_lookup(items)
    rows = read_daybook(input_path, date_by_ref_no)

    master_env, master_stats = build_masters(rows, items, masters_path)
    voucher_env, voucher_stats = build_vouchers(rows, items, vouchers_path)
    combine_xml(master_env, voucher_env, combined_path)

    LOGGER.info("Created masters XML: %s", masters_path)
    LOGGER.info("Created vouchers XML: %s", vouchers_path)
    LOGGER.info("Created combined XML: %s", combined_path)
    LOGGER.info("Master stats: %s", master_stats)
    LOGGER.info("Voucher stats: %s", voucher_stats)
    LOGGER.info("Recommended import order: 01_masters.xml first, then 02_vouchers.xml")


if __name__ == "__main__":
    main()
